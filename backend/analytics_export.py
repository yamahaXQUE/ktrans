"""Excel export for supervisor complaint analytics."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from backend.schemas import ComplaintAnalyticsDto


HEADERS = (
    "Дата звонка",
    "Дата анализа",
    "Дата отправки",
    "ID задачи Bitrix",
    "Оператор",
    "Отдел",
    "Тип задачи",
    "Основание",
    "Подтверждение жалобы",
    "Рекомендация модели",
    "Решение оператора",
    "Статус доставки",
    "Название задачи",
    "Описание",
    "Приоритет",
    "Ошибка доставки",
    "ID звонка",
)


def build_complaints_workbook(
    rows: list[dict[str, Any]],
    analytics: ComplaintAnalyticsDto,
) -> bytes:
    workbook = Workbook()
    complaints = workbook.active
    complaints.title = "Жалобы"
    complaints.freeze_panes = "A2"
    complaints.append(HEADERS)

    for row in rows:
        complaints.append(
            (
                _excel_datetime(row.get("call_started_at")),
                _excel_datetime(row.get("analyzed_at")),
                _excel_datetime(row.get("sent_at")),
                row.get("bitrix_item_id"),
                row.get("operator_name"),
                row.get("department"),
                row.get("task_type"),
                row.get("complaint_basis"),
                row.get("complaint_evidence"),
                "Создать" if row.get("should_create") else "Не создавать",
                row.get("review_decision") or "Ожидает решения",
                row.get("delivery_status") or "Не отправлялась",
                row.get("title"),
                row.get("description"),
                row.get("priority"),
                row.get("failure_reason"),
                str(row["call_id"]) if row.get("call_id") is not None else None,
            )
        )

    _style_header(complaints)
    complaints.auto_filter.ref = complaints.dimensions
    for column in ("A", "B", "C"):
        for cell in complaints[column][1:]:
            cell.number_format = "dd.mm.yyyy hh:mm"
    widths = (20, 20, 20, 18, 26, 28, 30, 28, 42, 22, 22, 22, 34, 60, 12, 48, 38)
    for index, width in enumerate(widths, start=1):
        complaints.column_dimensions[complaints.cell(1, index).column_letter].width = width
    for row in complaints.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = workbook.create_sheet("Аналитика")
    summary.merge_cells("A1:C1")
    summary["A1"] = "Воронка обработки звонков и жалоб"
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    summary["A1"].fill = PatternFill("solid", fgColor="7F1F93")
    summary["A1"].alignment = Alignment(vertical="center")
    summary["A2"] = "Сформировано"
    summary["B2"] = _excel_datetime(analytics.generated_at)
    summary["B2"].number_format = "dd.mm.yyyy hh:mm"
    summary["A3"] = "Период звонков"
    summary["B3"] = _period_label(analytics)
    summary.append(())
    summary.append(("Показатель", "Значение", "Пояснение"))
    metrics = (
        ("Всего звонков", analytics.total_calls, "Все загруженные звонки"),
        ("Расшифровано", analytics.analyzed_calls, "Анализ завершён"),
        (
            "Покрытие анализом",
            analytics.analysis_coverage_percent / 100,
            "Доля расшифрованных звонков",
        ),
        (
            "Ошибки анализа",
            analytics.analysis_failed_calls,
            "Расшифровка или анализ завершились ошибкой",
        ),
        ("Без результата", analytics.analysis_pending_calls, "Ожидают или обрабатываются"),
        ("Ручная очередь", analytics.manual_queue_calls, "Запрошены вручную и ещё не завершены"),
        (
            "Найдены жалобы",
            analytics.complaint_candidates,
            "Явные жалобы и негативная обратная связь",
        ),
        ("Подтверждено", analytics.confirmed_candidates, "Оператор подтвердил создание задачи"),
        ("Отклонено", analytics.rejected_candidates, "Оператор отклонил предложение"),
        ("Создано в Bitrix", analytics.total_complaints, "Успешно доставленные задачи"),
        ("Ошибки доставки", analytics.delivery_failed_tasks, "Bitrix не принял задачу"),
        (
            "Успешность доставки",
            analytics.delivery_success_percent / 100,
            "Создано от числа подтверждённых",
        ),
    )
    for metric in metrics:
        summary.append(metric)
    _style_row(summary, 5)
    summary["B8"].number_format = "0.0%"
    summary["B17"].number_format = "0.0%"
    summary.freeze_panes = "A6"
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 20
    summary.column_dimensions["C"].width = 48

    department_header_row = summary.max_row + 2
    summary.cell(department_header_row, 1, "Отдел")
    summary.cell(department_header_row, 2, "Созданные жалобы")
    summary.cell(department_header_row, 3, "Доля")
    _style_row(summary, department_header_row)
    for item in analytics.departments:
        summary.append((item.department, item.count, item.share_percent / 100))
        summary.cell(summary.max_row, 3).number_format = "0.0%"

    if analytics.departments:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Созданные жалобы по отделам"
        chart.y_axis.title = "Отдел"
        chart.x_axis.title = "Количество"
        chart.height = max(7, min(16, len(analytics.departments) * 0.65))
        chart.width = 17
        data = Reference(
            summary,
            min_col=2,
            min_row=department_header_row,
            max_row=department_header_row + len(analytics.departments),
        )
        categories = Reference(
            summary,
            min_col=1,
            min_row=department_header_row + 1,
            max_row=department_header_row + len(analytics.departments),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        summary.add_chart(chart, "E5")

    task_types = workbook.create_sheet("Типы задач")
    task_types.append(("Тип задачи", "Найдено жалоб", "Доля"))
    for item in analytics.task_types:
        task_types.append((item.task_type, item.count, item.share_percent / 100))
        task_types.cell(task_types.max_row, 3).number_format = "0.0%"
    _style_header(task_types)
    task_types.freeze_panes = "A2"
    task_types.auto_filter.ref = task_types.dimensions
    task_types.column_dimensions["A"].width = 38
    task_types.column_dimensions["B"].width = 18
    task_types.column_dimensions["C"].width = 14

    daily = workbook.create_sheet("Динамика")
    daily.append(
        (
            "Дата",
            "Звонки",
            "Расшифровано",
            "Ошибки анализа",
            "Найдены жалобы",
            "Создано задач",
        )
    )
    for item in analytics.daily:
        daily.append(
            (
                item.day,
                item.calls,
                item.analyzed_calls,
                item.analysis_failures,
                item.complaint_candidates,
                item.created_tasks,
            )
        )
        daily.cell(daily.max_row, 1).number_format = "dd.mm.yyyy"
    _style_header(daily)
    daily.freeze_panes = "A2"
    daily.auto_filter.ref = daily.dimensions
    daily.column_dimensions["A"].width = 14
    for column in ("B", "C", "D", "E", "F"):
        daily.column_dimensions[column].width = 19

    if analytics.daily:
        chart = LineChart()
        chart.style = 13
        chart.title = "Динамика обработки звонков"
        chart.y_axis.title = "Количество"
        chart.x_axis.title = "Дата"
        chart.height = 10
        chart.width = 22
        data = Reference(daily, min_col=2, max_col=6, min_row=1, max_row=len(analytics.daily) + 1)
        categories = Reference(daily, min_col=1, min_row=2, max_row=len(analytics.daily) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        daily.add_chart(chart, "H2")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _style_header(sheet: Any) -> None:
    _style_row(sheet, 1)


def _style_row(sheet: Any, row_number: int) -> None:
    fill = PatternFill("solid", fgColor="7F1F93")
    for cell in sheet[row_number]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _period_label(analytics: ComplaintAnalyticsDto) -> str:
    if not analytics.period_start or not analytics.period_end:
        return "Нет данных"
    return f"{analytics.period_start:%d.%m.%Y} — {analytics.period_end:%d.%m.%Y}"


def _excel_datetime(value: Any) -> Any:
    if value is None:
        return None
    if getattr(value, "tzinfo", None) is not None:
        return value.replace(tzinfo=None)
    return value


__all__ = ["HEADERS", "build_complaints_workbook"]
