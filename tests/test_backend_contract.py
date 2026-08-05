import os
import unittest
from types import SimpleNamespace
from datetime import date, datetime, timezone
from io import BytesIO
from unittest.mock import patch
from uuid import UUID

from backend.bitrix_auth import validate_current_user, validate_current_user_profile
from backend.router import _access_bypass_user_ids
from backend.analytics_export import build_complaints_workbook
from backend.main import app
from backend.schemas import (
    ComplaintAnalyticsDto,
    ComplaintDailyStatDto,
    ComplaintDepartmentStatDto,
    ComplaintTaskTypeStatDto,
    SessionUserDto,
    SourceCallDto,
)
from backend.sync_bitrix import mask_phone
from backend.task_delivery import (
    TaskDeliveryConfig,
    _bitrix_priority,
    _require_concrete_complaint,
    _task_description,
)
from bitrix import BitrixResponse


class FrontendContractTests(unittest.TestCase):
    def test_dtos_serialize_with_frontend_camel_case(self):
        operator_id = UUID("00000000-0000-0000-0000-000000000001")
        call = SourceCallDto(
            id=UUID("00000000-0000-0000-0000-000000000002"),
            statistic_id=99,
            operator_id=operator_id,
            operator_name="Иван Петров",
            direction="inbound",
            duration_seconds=30,
            started_at=datetime(2026, 7, 23, tzinfo=timezone.utc),
            phone_masked="+996 XX 1234",
            failed_code=None,
            transcript="Текст",
            conversation_title="Уточнение заказа",
            analysis_status="completed",
            analysis_requested=False,
            analysis_error=None,
        )

        payload = call.model_dump(mode="json")

        self.assertEqual(payload["statisticId"], 99)
        self.assertEqual(payload["operatorId"], str(operator_id))
        self.assertEqual(payload["conversationTitle"], "Уточнение заказа")
        self.assertNotIn("statistic_id", payload)

    def test_openapi_contains_frontend_routes(self):
        paths = app.openapi()["paths"]

        self.assertIn("/api/candidates", paths)
        self.assertIn("/api/candidates/{candidate_id}/confirm", paths)
        self.assertIn("/api/emulation/users", paths)
        self.assertIn("/api/emulation/session", paths)
        self.assertIn("/api/calls/{call_id}/analysis", paths)
        self.assertIn("/api/calls/{call_id}", paths)
        self.assertIn("/api/operators/{operator_id}/calls", paths)
        self.assertIn("/api/analytics/complaints", paths)
        self.assertIn("/api/analytics/complaints.xlsx", paths)
        query_names = {
            parameter["name"]
            for parameter in paths["/api/candidates"]["get"]["parameters"]
            if parameter["in"] == "query"
        }
        self.assertIn("operatorId", query_names)

    def test_session_role_contract(self):
        user = SessionUserDto(
            id=UUID("00000000-0000-0000-0000-000000000001"),
            display_name="Регина Лазаревич",
            work_position="Супервайзер",
            initials="РЛ",
            role="supervisor",
            department_ids=[82],
            source="bitrix",
        )

        self.assertEqual(user.model_dump(mode="json")["departmentIds"], [82])

    def test_excel_export_contains_complaints_and_analytics(self):
        from openpyxl import load_workbook

        analytics = ComplaintAnalyticsDto(
            total_complaints=1,
            total_calls=20,
            analyzed_calls=10,
            analysis_failed_calls=2,
            analysis_pending_calls=8,
            manual_queue_calls=1,
            complaint_candidates=3,
            confirmed_candidates=2,
            rejected_candidates=1,
            delivery_failed_tasks=1,
            analysis_coverage_percent=50,
            delivery_success_percent=50,
            period_start=datetime(2026, 7, 22, tzinfo=timezone.utc),
            period_end=datetime(2026, 7, 23, tzinfo=timezone.utc),
            generated_at=datetime(2026, 7, 23, tzinfo=timezone.utc),
            departments=[
                ComplaintDepartmentStatDto(
                    department="Поддержка",
                    count=1,
                    share_percent=100,
                )
            ],
            task_types=[
                ComplaintTaskTypeStatDto(
                    task_type="service_fm",
                    count=3,
                    share_percent=100,
                )
            ],
            daily=[
                ComplaintDailyStatDto(
                    day=date(2026, 7, 23),
                    calls=20,
                    analyzed_calls=10,
                    analysis_failures=2,
                    complaint_candidates=3,
                    created_tasks=1,
                )
            ],
        )
        content = build_complaints_workbook(
            [
                {
                    "call_started_at": datetime(2026, 7, 23, tzinfo=timezone.utc),
                    "analyzed_at": datetime(2026, 7, 23, tzinfo=timezone.utc),
                    "sent_at": datetime(2026, 7, 23, tzinfo=timezone.utc),
                    "bitrix_item_id": "321",
                    "operator_name": "Иван Петров",
                    "department": "Поддержка",
                    "task_type": "service_fm",
                    "complaint_basis": "explicit_complaint",
                    "complaint_evidence": "Клиент явно пожаловался.",
                    "should_create": True,
                    "review_decision": "confirmed",
                    "delivery_status": "created",
                    "title": "Разобрать жалобу",
                    "description": "Проверить обслуживание.",
                    "priority": 3,
                    "failure_reason": None,
                    "call_id": UUID("00000000-0000-0000-0000-000000000003"),
                }
            ],
            analytics,
        )

        self.assertTrue(content.startswith(b"PK"))
        workbook = load_workbook(BytesIO(content))
        self.assertEqual(
            workbook.sheetnames,
            ["Жалобы", "Аналитика", "Типы задач", "Динамика"],
        )
        self.assertEqual(workbook["Жалобы"]["E2"].value, "Иван Петров")
        self.assertEqual(workbook["Жалобы"]["L2"].value, "created")
        self.assertEqual(workbook["Аналитика"]["B6"].value, 20)
        self.assertEqual(workbook["Аналитика"]["B8"].value, 0.5)
        self.assertEqual(workbook["Типы задач"]["A2"].value, "service_fm")
        self.assertEqual(workbook["Динамика"]["B2"].value, 20)


class ConfigurationTests(unittest.TestCase):
    def test_native_task_default_department_is_configuration_driven(self):
        environment = {
            "BITRIX_TASK_DEFAULT_DEPARTMENT_ID": "82",
            "BITRIX_TASK_ADD_METHOD": "task.item.add",
        }
        with patch.dict(os.environ, environment, clear=False):
            config = TaskDeliveryConfig.from_env()

        self.assertEqual(config.default_department_id, 82)
        self.assertEqual(config.add_method, "task.item.add")

    def test_native_task_fields_preserve_routing_context(self):
        description = _task_description(
            {
                "description": "Проверить обращение.",
                "department": "Поддержка",
                "call_id": UUID(
                    "00000000-0000-0000-0000-000000000003"
                ),
            }
        )

        self.assertIn("Проверить обращение.", description)
        self.assertIn("Подразделение: Поддержка", description)
        self.assertIn("Источник: звонок", description)
        self.assertEqual(
            [_bitrix_priority(value) for value in range(1, 6)],
            ["0", "0", "1", "2", "2"],
        )

    def test_bitrix_delivery_requires_a_concrete_complaint(self):
        vague = SimpleNamespace(
            should_create=False,
            complaint_basis="explicit_negative_feedback",
            is_concrete_complaint=False,
            complaint_subject="",
            complaint_issue="",
        )
        with self.assertRaisesRegex(Exception, "concrete customer complaint"):
            _require_concrete_complaint(vague)

        concrete = SimpleNamespace(
            should_create=True,
            complaint_basis="explicit_complaint",
            is_concrete_complaint=True,
            complaint_subject="заказ 123",
            complaint_issue="привезли другой товар",
        )
        _require_concrete_complaint(concrete)

    def test_phone_is_masked_before_frontend(self):
        masked = mask_phone("+996 555 12 34 56")

        self.assertNotIn("555123456", masked.replace(" ", ""))
        self.assertTrue(masked.endswith("3456"))

    def test_iframe_user_is_verified_server_side(self):
        class FakeClient:
            def __init__(self, base_url, **kwargs):
                self.base_url = base_url
                self.kwargs = kwargs

            def call(self, method, params):
                self.method = method
                self.params = params
                return BitrixResponse(
                    result={
                        "ID": "10",
                        "NAME": "Регина",
                        "LAST_NAME": "Лазаревич",
                        "WORK_POSITION": "Супервайзер",
                    },
                    total=None,
                    next=None,
                    raw_response={"result": {"ID": "10"}},
                )

        with (
            patch("backend.bitrix_auth.BitrixClient", FakeClient),
            patch.dict(
                os.environ,
                {
                    "BITRIX_ALLOWED_PORTALS": "bitrix.kulikov.com",
                    "BITRIX_TLS_COMPATIBILITY": "true",
                },
                clear=False,
            ),
        ):
            user_id = validate_current_user(
                domain="bitrix.kulikov.com",
                access_token="temporary-token",
            )
            profile = validate_current_user_profile(
                domain="bitrix.kulikov.com",
                access_token="temporary-token",
            )

        self.assertEqual(user_id, 10)
        self.assertEqual(profile.display_name, "Регина Лазаревич")
        self.assertEqual(profile.work_position, "Супервайзер")

    def test_iframe_domain_allowlist_blocks_ssrf(self):
        with patch.dict(
            os.environ,
            {"BITRIX_ALLOWED_PORTALS": "bitrix.kulikov.com"},
            clear=False,
        ):
            with self.assertRaisesRegex(ValueError, "not allowed"):
                validate_current_user(
                    domain="localhost",
                    access_token="temporary-token",
                )

    def test_access_bypass_is_an_explicit_user_allowlist(self):
        with patch.dict(
            os.environ,
            {"BITRIX_ACCESS_BYPASS_USER_IDS": "255747, 10"},
            clear=False,
        ):
            self.assertEqual(_access_bypass_user_ids(), frozenset({255747, 10}))

        with patch.dict(
            os.environ,
            {"BITRIX_ACCESS_BYPASS_USER_IDS": "255747,invalid"},
            clear=False,
        ):
            with self.assertRaisesRegex(RuntimeError, "positive integers"):
                _access_bypass_user_ids()


if __name__ == "__main__":
    unittest.main()
